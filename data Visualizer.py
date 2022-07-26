from tkinter import messagebox
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import sqlite3
from tkinter import *
from tkinter import filedialog
from PIL import ImageTk, Image
from openpyxl.workbook import Workbook
from openpyxl import load_workbook



def open_filesp():
    global Entry1,Entry2,Entry3,Entry4,Entryt
    global entry1,entry2
    global root
    root.file_path = filedialog.askopenfilename(initialdir="Documents",title="Select A File",filetypes=[("Excel files", "*.xlsx")] ) 
    wb= load_workbook(root.file_path)
    ws= wb.active
    column_b = ws['B']
    Entry1=[]
    j=0
    for val in column_b:
        if(j!=0):
            Entry1.append(val.value)
        j=j+1
    column_c = ws['C']
    Entry2=[]
    i=0
    for val in column_c:
        if(i!=0):
            Entry2.append(val.value)
        i=i+1
    global r3
    r3 = Tk()
    r3.geometry("1920x1080")
    frame_a = LabelFrame(r3)
    frame_a.place(width=1800, height=550)  
    frame_a.pack(pady=20)
    
    
    Entry3 = Entry(frame_a, width=15, font=(" Maiandra GD", 15))
    Entry4 = Entry(frame_a, width=15, font=(" Maiandra GD", 15))
    Entryt = Entry(frame_a, width=20, font=(" Maiandra GD", 15))
    title  = Label(frame_a, text="Title:", padx=10, pady=5, font=(" Maiandra GD", 15))
    label3 = Label(frame_a, text="X Label: ", padx=10, font=(" Maiandra GD", 15))
    label4 = Label(frame_a, text="Y Label: ", padx=10, font=(" Maiandra GD", 15))
   
    submit = Button(frame_a, text="PLOT IT", bg="green", fg="white", width=10, height=3, command=plotscatterfile)
    back = Button(frame_a, text="<- BACK", height=1, width=10, command=new, font=("Consolas", 15), bg="grey",
                  fg="white")

    title.grid(row=0, column=0)
    Entryt.grid(row=0, column=1)

    label3.grid(row=1, column=0)
    Entry3.grid(row=1, column=1)

    label4.grid(row=1, column=2)
    Entry4.grid(row=1, column=3)

    submit.grid(row=3, column=1, pady=10)
    back.grid(row=3,column=2)
    entry1=""
    entry2=""
    for i in Entry1:
        entry1=entry1+str(Entry1)
        if(i!=len(Entry1)):
            entry1=entry1+","
    for j in Entry2:
        entry2=entry2+str(Entry2)
        if(i!=len(Entry2)):
            entry2=entry2+","
    
    
    r3.mainloop()

def plotscatterfile():
    # frame for plot
    panel = LabelFrame(r3)
    panel.pack()

    if len(Entry1) == 0 or len(Entry2) == 0 or len(Entry3.get()) == 0 or len(Entry4.get()) == 0 or len(
            Entryt.get()) == 0:
        messagebox.showinfo("Error", "Please Enter a Valid input")
    else:
        # getting input
        try:
            x = Entry1
            y = Entry2
        except:
            messagebox.showinfo("Error", "Enter Comma seperated numbers only")
        else:
            # variables
            type = "scatter"
            title = Entryt.get()
            xlabel = Entry3.get()
            ylabel = Entry4.get()
            param0 = """select title from plots;"""
            titles = conn.execute(param0)
            titles_lis = []
            for r in titles:
                titles_lis.append(r[0])
            if title not in titles_lis:
                # inserting values in database
                param = """insert into plots(title, type, x_values, y_values, x_label, y_label) values(?, ?, ?, ?, ?, ?);"""
                columns = (title, type, entry1, entry2, xlabel, ylabel)
                cursor.execute(param, columns)
                conn.commit()

                fig = Figure(figsize=(9,4), dpi=100)
                plot1 = fig.add_subplot(111)
                plot1.set_xlabel(Entry3.get())
                plot1.set_ylabel(Entry4.get())
                plot1.scatter(x, y)
                name = "scatter" + title + ".png"
                fig.savefig("C:/Users/Hemanth.M/Desktop/project images/" + name)
                canvas = FigureCanvasTkAgg(fig, master=panel)
                canvas.draw()
                canvas.get_tk_widget().pack()
                path = "C:/Users/Hemanth.M/Desktop/project images/" + name
                param = """update plots set plot=? where title=?;"""
                column = (path, title)
                cursor.execute(param, column)
                conn.commit()
            else:
                messagebox.showinfo("Error", "Please Choose an unique Title.")
    
 
def open_filesh():
    global Entry1,Entry3,Entryt
    global entry1
    global root
    root.file_path = filedialog.askopenfilename(initialdir="Documents",title="Select A File",filetypes=[("Excel files", "*.xlsx")] ) 
    wb= load_workbook(root.file_path)
    ws= wb.active
    column_a = ws['A']
    Entry1=[]
    j=0
    for val in column_a:
        if(j!=0):
            Entry1.append(val.value)
        j=j+1
    column_b = ws['B']
    global r3
    r3 = Tk()
    r3.geometry("1920x1080")
    frame_a = LabelFrame(r3)
    frame_a.place(width=1800, height=550)  
    frame_a.pack(pady=20)
    
    
    Entry3 = Entry(frame_a, width=15, font=(" Maiandra GD", 15))
    Entryt = Entry(frame_a, width=20, font=(" Maiandra GD", 15))
    title  = Label(frame_a, text="Title:", padx=5, pady=5, font=(" Maiandra GD", 15))
    label2 = Label(frame_a, text="Label: ", padx=5,pady=3, font=(" Maiandra GD", 15))
   
    submit = Button(frame_a, text="PLOT IT", bg="green", fg="white", width=10, height=3, command=histfile)
    back = Button(frame_a, text="<- BACK", height=1, width=10, command=new, font=("Consolas", 15), bg="grey",
                  fg="white")

    title.grid(row=0, column=0)
    Entryt.grid(row=0, column=1,padx=5)


    label2.grid(row=1, column=0)
    Entry3.grid(row=1, column=1,)

    submit.grid(row=2, column=1, pady=10)
    back.grid(row=2,column=0,padx=5)
    entry1=""
    for i in Entry1:
        entry1=entry1+str(Entry1)
        if(i!=len(Entry1)):
            entry1=entry1+","
    
    r3.mainloop()
def histfile():
    panel = LabelFrame(r3, text="PLOT")
    panel.pack()

    # getting input
    if len(Entryt.get()) == 0 or len(Entry1) == 0 or len(Entry3.get()) == 0:
        messagebox.showinfo("No Input Found", "Please fill all the information")
    else:
        try:
            x = Entry1
        except:
            messagebox.showinfo("Invalid Input", "please enter comma seperated Integers")
        else:
            # setting variables
            type = "hist"
            xlabel = entry3.get()
            ylabel = "frequency"
            title = Entryt.get()
            param0 = """select title from plots;"""
            titles = conn.execute(param0)
            titles_lis = []
            for r in titles:
                titles_lis.append(r[0])
            if title not in titles_lis:
                counted = {}
                for v in x:
                    if v not in counted.keys():
                        counted[v] = 1
                    else:
                        counted[v] += 1
                y = ','.join(str(x) for x in counted.values())

                param = "insert into plots (title, type, x_values,y_values, x_label, y_label) values(?, ?, ?, ?, ?, ?);"
                columns = (title, type, entry1, y, xlabel, ylabel)
                cursor.execute(param, columns)
                conn.commit()

                fig = Figure(figsize=(9, 4.5), dpi=100)
                plot1 = fig.add_subplot(111)
                plot1.set_xlabel(entry3.get())
                plot1.hist(x, bins=20)
                name = "hist" + title + ".png"
                fig.savefig("C:/Users/Hemanth.M/Desktop/project images/" + name)
                canvas = FigureCanvasTkAgg(fig, master=panel)
                canvas.draw()
                canvas.get_tk_widget().pack()

                path = "C:/Users/Hemanth.M/Desktop/project images/" + name
                param = """update plots set plot=? where title=?;"""
                column = (path, title)

                cursor.execute(param, column)
                conn.commit()

def open_filep():
    global Entry1,Entry3,Entryt
    global entry1,entry3
    global root
    root.file_path = filedialog.askopenfilename(initialdir="Documents",title="Select A File",filetypes=[("Excel files", "*.xlsx")] ) 
    wb= load_workbook(root.file_path)
    ws= wb.active
    column_a = ws['A'] #lables
    column_b = ws['B'] #values
    Entry3=[] #lables
    Entry1=[] #val
    j=0
    for val in column_a:
        if(j!=0):
            Entry3.append(val.value)
        j=j+1
    i=0
    for val in column_b:
        if(i!=0):
            Entry1.append(val.value)
        i=i+1
    global r3
    r3 = Tk()
    r3.geometry("1920x1080")
    frame_a = LabelFrame(r3)
    frame_a.place(width=1800, height=550)  
    frame_a.pack(pady=20)
    
    Entryt = Entry(frame_a, width=20, font=(" Maiandra GD", 15))
    title  = Label(frame_a, text="Title:", padx=5, pady=5, font=(" Maiandra GD", 15))
   
    submit = Button(frame_a, text="PLOT IT", bg="green", fg="white", width=10, height=3, command=piefile)
    back = Button(frame_a, text="<- BACK", height=1, width=10, command=new, font=("Consolas", 15), bg="grey",
                  fg="white")

    title.grid(row=0, column=0)
    Entryt.grid(row=0, column=1,padx=5)

    submit.grid(row=2, column=1, pady=10)
    back.grid(row=2,column=0,padx=5)
    entry1=""
    entry3=""
    for i in Entry1:
        entry1=entry1+str(Entry1)
        if(i!=len(Entry1)):
            entry1=entry1+","
    for j in Entry3:
        entry3=entry3+str(Entry3)
        if(i!=len(Entry3)):
            entry3=entry3+","
    r3.mainloop()
    
    
def piefile():
    panel = LabelFrame(r3, text="PLOT")
    panel.pack()

    if len(Entryt.get()) == 0 or len(Entry1) == 0 or len(Entry3) == 0:
        messagebox.showinfo("No Input Found", "Please fill all the information")
    else:
        try:
            # getting input
            x =Entry1
            lbls = Entry3
        except:
            messagebox.showinfo("Invalid Input", "Please enter comma seperated numbers")
        else:
            title = Entryt.get()
            pielabel = entry3

            param0 = """select title from pieplots;"""
            titles = conn.execute(param0)
            titles_lis = []
            for r in titles:
                titles_lis.append(r[0])
            if title not in titles_lis:

                param = """insert into pieplots(title,pie_values,labels) values(?, ?, ?);"""
                columns = (title, entry1, pielabel)
                cursor.execute(param, columns)
                conn.commit()

                fig = Figure(figsize=(9,4.5), dpi=100)
                plot1 = fig.add_subplot(111)
                plot1.pie(x, labels=lbls)
                name = "pie" + title + ".png"
                fig.savefig("C:/Users/Hemanth.M/Desktop/project images/" + name)
                canvas = FigureCanvasTkAgg(fig, master=panel)
                canvas.draw()
                canvas.get_tk_widget().pack()

                path = "C:/Users/Hemanth.M/Desktop/project images/" + name
                param = """update pieplots set path=? where title=?;"""
                column = (path, title)

                cursor.execute(param, column)
                conn.commit()
            else:
                messagebox.showinfo("Error", "Please enter Unique Title")
def open_fileb():
    global Entry1,Entry2,Entry3,Entry4,Entryt
    global entry1,entry2
    global root
    root.file_path = filedialog.askopenfilename(initialdir="Documents",title="Select A File",filetypes=[("Excel files", "*.xlsx")] ) 
    wb= load_workbook(root.file_path)
    ws= wb.active
    column_b = ws['B']
    Entry1=[]
    j=0
    for val in column_b:
        if(j!=0):
            Entry1.append(val.value)
        j=j+1
    column_c = ws['C']
    Entry2=[]
    i=0
    for val in column_c:
        if(i!=0):
            Entry2.append(val.value)
        i=i+1
    global r3
    r3 = Tk()
    r3.geometry("1920x1080")
    frame_a = LabelFrame(r3)
    frame_a.place(width=1800, height=550)  
    frame_a.pack(pady=20)
    
    
    Entry3 = Entry(frame_a, width=15, font=(" Maiandra GD", 15))
    Entry4 = Entry(frame_a, width=15, font=(" Maiandra GD", 15))
    Entryt = Entry(frame_a, width=20, font=(" Maiandra GD", 15))
    title  = Label(frame_a, text="Title:", padx=10, pady=5, font=(" Maiandra GD", 15))
    label3 = Label(frame_a, text="X Label: ", padx=10, font=(" Maiandra GD", 15))
    label4 = Label(frame_a, text="Y Label: ", padx=10, font=(" Maiandra GD", 15))
   
    submit = Button(frame_a, text="PLOT IT", bg="green", fg="white", width=10, height=3, command=barfile)
    back = Button(frame_a, text="<- BACK", height=1, width=10, command=new, font=("Consolas", 15), bg="grey",
                  fg="white")

    title.grid(row=0, column=0)
    Entryt.grid(row=0, column=1)

    label3.grid(row=1, column=0)
    Entry3.grid(row=1, column=1)

    label4.grid(row=1, column=2)
    Entry4.grid(row=1, column=3)

    submit.grid(row=3, column=1, pady=10)
    back.grid(row=3,column=2)
    entry1=""
    entry2=""
    for i in Entry1:
        entry1=entry1+str(Entry1)
        if(i!=len(Entry1)):
            entry1=entry1+","
    for j in Entry2:
        entry2=entry2+str(Entry2)
        if(i!=len(Entry2)):
            entry2=entry2+","

def barfile():
    panel = LabelFrame(r3, text="PLOT")
    panel.pack()

    if len(Entryt.get()) == 0 or len(Entry1) == 0 or len(Entry2) == 0 or len(Entry3.get()) == 0 or len(
            Entry4.get()) == 0:
        messagebox.showinfo("No Input Found", "Please fill all the information")
    else:
        try:
            # getting input
            x = Entry1
            y = Entry2
        except:
            messagebox.showinfo("Invalid Input", "please enter comma seperated Integers")
        else:
            # setting variables
            type = "bar"
            xlabel = Entry3.get()
            ylabel = Entry4.get()
            title = Entryt.get()
            param0 = """select title from plots;"""
            titles = conn.execute(param0)
            titles_lis = []
            for r in titles:
                titles_lis.append(r[0])
            if title not in titles_lis:
                param = """insert into plots(title,type,x_values,y_values,x_label,y_label) values(?,?,?,?,?,?);"""
                columns = (title, type, entry1, entry2, xlabel, ylabel)
                cursor.execute(param, columns)
                conn.commit()

                fig = Figure(figsize=(10,4), dpi=100)
                plot1 = fig.add_subplot(111)
                plot1.set_xlabel(entry3.get())
                plot1.set_ylabel(entry4.get())
                plot1.bar(x, y)
                name = "bar" + title + ".png"
                fig.savefig("C:/Users/Hemanth.M/Desktop/project images/" + name)
                canvas = FigureCanvasTkAgg(fig, master=panel)
                canvas.draw()
                canvas.get_tk_widget().pack()

                path = "C:/Users/Hemanth.M/Desktop/project images/" + name
                param = """update plots set plot=? where title=?;"""
                column = (path, title)

                cursor.execute(param, column)
                conn.commit()
            else:
                messagebox.showinfo("Error", "Please enter Unique Title")



def open_filel():
    global Entry1,Entry2,Entry3,Entry4,Entryt
    global entry1,entry2
    global root
    root.file_path = filedialog.askopenfilename(initialdir="Documents",title="Select A File",filetypes=[("Excel files", "*.xlsx")] ) 
    wb= load_workbook(root.file_path)
    ws= wb.active
    column_b = ws['B']
    Entry1=[]
    j=0
    for val in column_b:
        if(j!=0):
            Entry1.append(val.value)
        j=j+1
    column_c = ws['C']
    Entry2=[]
    i=0
    for val in column_c:
        if(i!=0):
            Entry2.append(val.value)
        i=i+1
    global r3
    r3 = Tk()
    r3.geometry("1920x1080")
    frame_a = LabelFrame(r3)
    frame_a.place(width=1800, height=550)  
    frame_a.pack(pady=20)
    
    
    Entry3 = Entry(frame_a, width=15, font=(" Maiandra GD", 15))
    Entry4 = Entry(frame_a, width=15, font=(" Maiandra GD", 15))
    Entryt = Entry(frame_a, width=20, font=(" Maiandra GD", 15))
    title  = Label(frame_a, text="Title:", padx=10, pady=5, font=(" Maiandra GD", 15))
    label3 = Label(frame_a, text="X Label: ", padx=10, font=(" Maiandra GD", 15))
    label4 = Label(frame_a, text="Y Label: ", padx=10, font=(" Maiandra GD", 15))
   
    submit = Button(frame_a, text="PLOT IT", bg="green", fg="white", width=10, height=3, command=linefile)
    back = Button(frame_a, text="<- BACK", height=1, width=10, command=new, font=("Consolas", 15), bg="grey",
                  fg="white")

    title.grid(row=0, column=0)
    Entryt.grid(row=0, column=1)

    label3.grid(row=1, column=0)
    Entry3.grid(row=1, column=1)

    label4.grid(row=1, column=2)
    Entry4.grid(row=1, column=3)

    submit.grid(row=3, column=1, pady=10)
    back.grid(row=3,column=2)
    entry1=""
    entry2=""
    for i in Entry1:
        entry1=entry1+str(Entry1)
        if(i!=len(Entry1)):
            entry1=entry1+","
    for j in Entry2:
        entry2=entry2+str(Entry2)
        if(i!=len(Entry2)):
            entry2=entry2+","
    
    
    r3.mainloop()

def linefile():
    panel = LabelFrame(r3, text="PLOT")
    panel.pack()

    if len(Entryt.get()) == 0 or len(Entry1) == 0 or len(Entry2) == 0 or len(Entry3.get()) == 0 or len(
            Entry4.get()) == 0:
        messagebox.showinfo("No Input Found", "Please fill all the information")
    else:
        try:
            # getting input
            x = Entry1
            y = Entry2
        except:
            messagebox.showinfo("Invalid Input", "please enter comma seperated Integers")
        else:
            # setting variables
            type = "line"
            xlabel = Entry3.get()
            ylabel = Entry4.get()
            title = Entryt.get()
            param0 = """select title from plots;"""
            titles = conn.execute(param0)
            titles_lis = []
            for r in titles:
                titles_lis.append(r[0])
            if title not in titles_lis:
                param = """insert into plots(title,type,x_values,y_values,x_label,y_label) values(?,?,?,?,?,?);"""
                columns = (title, type, entry1, entry2, xlabel, ylabel)
                cursor.execute(param, columns)
                conn.commit()

                fig = Figure(figsize=(10,4), dpi=100)
                plot1 = fig.add_subplot(111)
                plot1.set_xlabel(Entry3.get())
                plot1.set_ylabel(Entry4.get())
                plot1.plot(x, y)
                name = "line" + title + ".png"
                fig.savefig("C:/Users/Hemanth.M/Desktop/project images/" + name)
                canvas = FigureCanvasTkAgg(fig, master=panel)
                canvas.draw()
                canvas.get_tk_widget().pack()

                path = "C:/Users/Hemanth.M/Desktop/project images/" + name
                param = """update plots set plot=? where title=?;"""
                column = (path, title)

                cursor.execute(param, column)
                conn.commit()
            else:
                messagebox.showinfo("Error", "Please enter Unique Title")




    
    

def scatterReset():
    r1.destroy()
    scatterplt(True)


def plotscatter():
    # frame for plot
    panel = LabelFrame(r1)
    panel.pack()

    if len(entry1.get()) == 0 or len(entry2.get()) == 0 or len(entry3.get()) == 0 or len(entry4.get()) == 0 or len(
            entryt.get()) == 0:
        messagebox.showinfo("Error", "Please Enter a Valid input")
    else:
        # getting input
        try:
            x = list(map(int, entry1.get().split(",")))
            y = list(map(int, entry2.get().split(",")))
        except:
            messagebox.showinfo("Error", "Enter Comma seperated numbers only")
        else:
            # variables
            type = "scatter"
            title = entryt.get()
            xlabel = entry3.get()
            ylabel = entry4.get()
            param0 = """select title from plots;"""
            titles = conn.execute(param0)
            titles_lis = []
            for r in titles:
                titles_lis.append(r[0])
            if title not in titles_lis:
                # inserting values in database
                param = """insert into plots(title, type, x_values, y_values, x_label, y_label) values(?, ?, ?, ?, ?, ?);"""
                columns = (title, type, entry1.get(), entry2.get(), xlabel, ylabel)
                cursor.execute(param, columns)
                conn.commit()
                fig = Figure(figsize=(10,4), dpi=100)
                plot1 = fig.add_subplot(111)
                plot1.set_xlabel(entry3.get())
                plot1.set_ylabel(entry4.get())
                plot1.scatter(x, y)
                name = "scatter" + title + ".png"
                fig.savefig("C:/Users/Hemanth.M/Desktop/project images/" + name)
                canvas = FigureCanvasTkAgg(fig, master=panel)
                canvas.draw()
                canvas.get_tk_widget().pack()
                path = "C:/Users/Hemanth.M/Desktop/project images/" + name
                param = """update plots set plot=? where title=?;"""
                column = (path, title)
                cursor.execute(param, column)
                conn.commit()
            else:
                messagebox.showinfo("Error", "Please Choose an unique Title.")


def scatterplt(destroyed=False):
    # destroying old window
    global root
    if not destroyed:
        root.destroy()

    # creating new window
    global r1
    r1 = Tk()
    r1.geometry("1920x1080")

   
    frame_2 = LabelFrame(r1)
    frame_2.place(width=1800, height=550)  
    frame_2.pack(pady=20)
   
    # creating labels for input fields
    label1 = Label(frame_2, text="Enter X Values:", padx=5, font=(" Maiandra GD", 15))
    label2 = Label(frame_2, text="Enter Y Values:", padx=5, font=(" Maiandra GD", 15))
    label3 = Label(frame_2, text="X Label: ", padx=5, font=(" Maiandra GD", 15))
    label4 = Label(frame_2, text="Y Label: ", padx=5, font=(" Maiandra GD", 15))
    title = Label(frame_2, text="Title:", padx=5, pady=3, font=(" Maiandra GD", 15))

    # creating input fields
    global entry1, entry2, entry3, entry4, entryt
    entry1 = Entry(frame_2, width=30, font=(" Maiandra GD", 15))
    entry2 = Entry(frame_2, width=30, font=(" Maiandra GD", 15))
    entry3 = Entry(frame_2, width=15, font=(" Maiandra GD", 15))
    entry4 = Entry(frame_2, width=15, font=(" Maiandra GD", 15))
    entryt = Entry(frame_2, width=30, font=(" Maiandra GD", 15))

    # creating submit button to frame
    submit = Button(frame_2, text="PLOT IT", bg="green", fg="white", width=20, height=3, command=plotscatter)
    
    #creating select file button
    s_file = Button(frame_2,text ="select from file",bg= "blue",fg= "white",width =20 ,height = 3,command =open_filesp )

    # creating back button to window
    back = Button(frame_2, text="<- BACK", height=1, width=10, command=new, font=("Consolas", 15), bg="grey",
                  fg="white")
    back.grid(row=3, column=0)

    # creating reset button to frame
    reset = Button(frame_2, text="⟲", font=("Consolas", 20), width=10,height=1, bg="#DC8474", fg="white", command=scatterReset)
    reset.grid(row=3, column=2)

    # adding text fields,labels and buttons in order
    title.grid(row=0, column=0)
    entryt.grid(row=0, column=1)

    label1.grid(row=1, column=0)
    entry1.grid(row=1, column=1, padx=20, pady=10)

    label3.grid(row=1, column=2)
    entry3.grid(row=1, column=3, padx=10)

    label2.grid(row=2, column=0)
    entry2.grid(row=2, column=1, padx=20, pady=10)

    label4.grid(row=2, column=2)
    entry4.grid(row=2, column=3, padx=10)

    submit.grid(row=3, column=1, pady=20)
    
    s_file.grid(row=3,column=3)

    r1.mainloop()


def histReset():
    r1.destroy()
    histplt(True)


def plothist():
    # frame for plot
    panel = LabelFrame(r1, text="PLOT")
    panel.pack()

    # getting input
    if len(entryt.get()) == 0 or len(entry1.get()) == 0 or len(entry3.get()) == 0:
        messagebox.showinfo("No Input Found", "Please fill all the information")
    else:
        try:
            x = list(map(int, entry1.get().split(",")))
        except:
            messagebox.showinfo("Invalid Input", "please enter comma seperated Integers")
        else:
            # setting variables
            type = "hist"
            xlabel = entry3.get()
            ylabel = "frequency"
            title = entryt.get()
            param0 = """select title from plots;"""
            titles = conn.execute(param0)
            titles_lis = []
            for r in titles:
                titles_lis.append(r[0])
            if title not in titles_lis:
                counted = {}
                for v in x:
                    if v not in counted.keys():
                        counted[v] = 1
                    else:
                        counted[v] += 1
                y = ','.join(str(x) for x in counted.values())

                param = "insert into plots (title, type, x_values,y_values, x_label, y_label) values(?, ?, ?, ?, ?, ?);"
                columns = (title, type, entry1.get(), y, xlabel, ylabel)
                cursor.execute(param, columns)
                conn.commit()

                fig = Figure(figsize=(10, 5), dpi=100)
                plot1 = fig.add_subplot(111)
                plot1.set_xlabel(entry3.get())
                plot1.hist(x, bins=20)
                name = "hist" + title + ".png"
                fig.savefig("C:/Users/Hemanth.M/Desktop/project images/" + name)
                canvas = FigureCanvasTkAgg(fig, master=panel)
                canvas.draw()
                canvas.get_tk_widget().pack()

                path = "C:/Users/Hemanth.M/Desktop/project images/" + name
                param = """update plots set plot=? where title=?;"""
                column = (path, title)

                cursor.execute(param, column)
                conn.commit()


def histplt(destroyed=False):
    # destroying old window
    global root
    if not destroyed:
        root.destroy()

    # creating new window
    global r1
    r1 = Tk()
    r1.geometry("1920x1080")

    # adding frame to window
    frame_2 = LabelFrame(r1)
    frame_2.place(width=1920, height=480)
    frame_2.pack()

    # creating labels for input fields
    label1 = Label(frame_2, text="Enter X Values :", padx=20, font=(" Maiandra GD", 15))
    label3 = Label(frame_2, text="Label:", padx=20, font=(" Maiandra GD", 15))
    title = Label(frame_2, text="Title:", padx=20, font=("Maiandra GD", 15))

    # creating input fields
    global entry1, entry3, entryt
    entry1 = Entry(frame_2, width=30, font=(" Maiandra GD", 15))
    entry3 = Entry(frame_2, width=15, font=(" Maiandra GD", 15))
    entryt = Entry(frame_2, width=30, font=(" Maiandra GD", 15))

    # creating submit button to frame
    submit = Button(frame_2, text="PLOT IT", bg="green", fg="white", width=20, height=3, command=plothist)

    # creating back button to window
    back = Button(frame_2, text="<- BACK", height=1, width=10, command=new, font=("Consolas", 15), bg="grey",
                  fg="white")
    back.grid(row=2, column=0)

    # creating reset button to frame
    reset = Button(frame_2, text="⟲", font=("Consolas", 20), width=8, bg="#DC8474", fg="white", command=histReset)
    reset.grid(row=2, column=2)
    
    #creating select file button
    s_file = Button(frame_2,text ="select from file",bg= "blue",fg= "white",width =20 ,height = 3,command =open_filesh)
    
    # adding text fields,labels and buttons in order
    title.grid(row=0, column=0)
    entryt.grid(row=0, column=1)

    label1.grid(row=1, column=0)
    entry1.grid(row=1, column=1, padx=20, pady=10)

    label3.grid(row=1, column=2)
    entry3.grid(row=1, column=3, padx=10)

    submit.grid(row=2, column=1, pady=20)
    
    s_file.grid(row=2,column=3)


    r1.mainloop()


def pieReset():
    r1.destroy()
    pieplt(True)


def plotpie():
    panel = LabelFrame(r1, text="PLOT")
    panel.pack()

    if len(entryt.get()) == 0 or len(entry1.get()) == 0 or len(entry3.get()) == 0:
        messagebox.showinfo("No Input Found", "Please fill all the information")
    else:
        try:
            # getting input
            x = list(map(int, entry1.get().split(",")))
            lbls = list(entry3.get().split(","))
        except:
            messagebox.showinfo("Invalid Input", "Please enter comma seperated numbers")
        else:
            title = entryt.get()
            pielabel = entry3.get()

            param0 = """select title from pieplots;"""
            titles = conn.execute(param0)
            titles_lis = []
            for r in titles:
                titles_lis.append(r[0])
            if title not in titles_lis:

                param = """insert into pieplots(title,pie_values,labels) values(?, ?, ?);"""
                columns = (title, entry1.get(), pielabel)
                cursor.execute(param, columns)
                conn.commit()

                fig = Figure(figsize=(10,5), dpi=100)
                plot1 = fig.add_subplot(111)
                plot1.pie(x, labels=lbls)
                name = "pie" + title + ".png"
                fig.savefig("C:/Users/Hemanth.M/Desktop/project images/" + name)
                canvas = FigureCanvasTkAgg(fig, master=panel)
                canvas.draw()
                canvas.get_tk_widget().pack()

                path = "C:/Users/Hemanth.M/Desktop/project images/" + name
                param = """update pieplots set path=? where title=?;"""
                column = (path, title)

                cursor.execute(param, column)
                conn.commit()
            else:
                messagebox.showinfo("Error", "Please enter Unique Title")


def pieplt(destroyed=False):
    # destroying old window
    global root
    if not destroyed:
        root.destroy()

    # creating new window
    global r1
    r1 = Tk()
    r1.geometry("1920x1080")

    # adding frame to window
    frame_2 = LabelFrame(r1)
    frame_2.place(width=1800, height=480)
    frame_2.pack()

    # creating labels for input fields
    label1 = Label(frame_2, text="Enter Values:", padx=20, font=(" Maiandra GD", 15))
    label3 = Label(frame_2, text="Labels:", padx=20, font=(" Maiandra GD", 15))
    title = Label(frame_2, text="Title:", padx=20, font=(" Maiandra GD", 15))

    # creating input fields
    global entry1, entry3, entryt
    entry1 = Entry(frame_2, width=30, font=(" Maiandra GD", 15))
    entry3 = Entry(frame_2, width=15, font=(" Maiandra GD", 15))
    entryt = Entry(frame_2, width=30, font=(" Maiandra GD", 15))

    # creating submit button to frame
    submit = Button(frame_2, text="PLOT IT", bg="green", fg="white", width=20, height=3, command=plotpie)

    # creating back button to window
    back = Button(frame_2, text="<- BACK", height=1, width=10, command=new, font=("Consolas", 15), bg="grey",
                  fg="white")
    back.grid(row=2, column=0)
    #creating select file button
    s_file = Button(frame_2,text ="select from file",bg= "blue",fg= "white",width =20 ,height = 3,command =open_filep)
    

    # creating reset button to frame
    reset = Button(frame_2, text="⟲", font=("Consolas", 20), width=8, bg="#DC8474", fg="white", command=pieReset)
    reset.grid(row=2, column=2)

    # adding text fields,labels and buttons in order
    title.grid(row=0, column=0)
    entryt.grid(row=0, column=1)

    label1.grid(row=1, column=0)
    entry1.grid(row=1, column=1, padx=20, pady=10)

    label3.grid(row=1, column=2)
    entry3.grid(row=1, column=3, padx=10)

    submit.grid(row=2, column=1, pady=20)
    
    s_file.grid(row=2,column=3)
    

    r1.mainloop()


def lineReset():
    r1.destroy()
    lineplt(True)


def plotline():
    # frame for plot
    panel = LabelFrame(r1, text="PLOT")
    panel.pack()

    if len(entryt.get()) == 0 or len(entry1.get()) == 0 or len(entry2.get()) == 0 or len(entry3.get()) == 0 or len(
            entry4.get()) == 0:
        messagebox.showinfo("No Input Found", "Please fill all the information")
    else:
        try:
            # getting input
            x = list(map(int, entry1.get().split(",")))
            y = list(map(int, entry2.get().split(",")))
        except:
            messagebox.showinfo("Invalid Input", "please enter comma seperated Integers")
        else:
            # setting variables
            type = "line"
            xlabel = entry3.get()
            ylabel = entry4.get()
            title = entryt.get()
            param0 = """select title from plots;"""
            titles = conn.execute(param0)
            titles_lis = []
            for r in titles:
                titles_lis.append(r[0])
            if title not in titles_lis:
                param = """insert into plots(title,type,x_values,y_values,x_label,y_label) values(?,?,?,?,?,?);"""
                columns = (title, type, entry1.get(), entry2.get(), xlabel, ylabel)
                cursor.execute(param, columns)
                conn.commit()

                fig = Figure(figsize=(10,4), dpi=100)
                plot1 = fig.add_subplot(111)
                plot1.set_xlabel(entry3.get())
                plot1.set_ylabel(entry4.get())
                plot1.plot(x, y)
                name = "line" + title + ".png"
                fig.savefig("C:/Users/Hemanth.M/Desktop/project images/" + name)
                canvas = FigureCanvasTkAgg(fig, master=panel)
                canvas.draw()
                canvas.get_tk_widget().pack()

                path = "C:/Users/Hemanth.M/Desktop/project images/" + name
                param = """update plots set plot=? where title=?;"""
                column = (path, title)

                cursor.execute(param, column)
                conn.commit()
            else:
                messagebox.showinfo("Error", "Please enter Unique Title")


def lineplt(destroyed=False):
    # destroying old window
    global root
    if not destroyed:
        root.destroy()

    # creating new window
    global r1
    r1 = Tk()
    r1.geometry("1920x1080")

    # adding frame to window
    frame_2 = LabelFrame(r1)
    frame_2.place(width=1750, height=480)
    frame_2.pack()

    # creating labels for input fields
    label1 = Label(frame_2, text="Enter X Values:", padx=20, font=(" Maiandra GD", 15))
    label2 = Label(frame_2, text="Enter Y Values:", padx=20, font=(" Maiandra GD", 15))
    label3 = Label(frame_2, text="X Label: ", padx=20, font=(" Maiandra GD", 15))
    label4 = Label(frame_2, text="Y Label:", padx=20, font=(" Maiandra GD", 15))
    title = Label(frame_2, text="Title: ", padx=20, font=(" Maiandra GD", 15))

    # creating input fields
    global entry1, entry2, entry3, entry4, entryt
    entry1 = Entry(frame_2, width=30, font=(" Maiandra GD", 15))
    entry2 = Entry(frame_2, width=30, font=(" Maiandra GD", 15))
    entry3 = Entry(frame_2, width=15, font=(" Maiandra GD", 15))
    entry4 = Entry(frame_2, width=15, font=(" Maiandra GD", 15))
    entryt = Entry(frame_2, width=30, font=(" Maiandra GD", 15))

    # creating submit button to frame
    submit = Button(frame_2, text="PLOT IT", bg="green", fg="white", width=20, height=3, command=plotline)

    # creating back button to window
    back = Button(frame_2, text="<- BACK", height=1, width=10, command=new, font=("Consolas", 15), bg="grey",
                  fg="white")
    back.grid(row=3, column=0)
    #creating select file button
    s_file = Button(frame_2,text ="selct from file",bg= "blue",fg= "white",width =20 ,height = 3,command =open_filel)
    

    # creating reset button to frame
    reset = Button(frame_2, text="⟲", font=("Consolas", 20), width=8, bg="#DC8474", fg="white", command=lineReset)
    reset.grid(row=3, column=2)

    # adding text fields,labels and buttons in order
    title.grid(row=0, column=0)
    entryt.grid(row=0, column=1)

    label1.grid(row=1, column=0)
    entry1.grid(row=1, column=1, padx=20, pady=10)

    label3.grid(row=1, column=2)
    entry3.grid(row=1, column=3, padx=10)

    label2.grid(row=2, column=0)
    entry2.grid(row=2, column=1, padx=20, pady=10)

    label4.grid(row=2, column=2)
    entry4.grid(row=2, column=3, padx=10)

    submit.grid(row=3, column=1, pady=20)
    s_file.grid(row=3,column=3)

    r1.mainloop()


def barReset():
    r1.destroy()
    barplt(True)


def plotbar():
    # frame for plot
    panel = LabelFrame(r1, text="PLOT")
    panel.pack()

    if len(entryt.get()) == 0 or len(entry1.get()) == 0 or len(entry2.get()) == 0 or len(entry3.get()) == 0 or len(
            entry4.get()) == 0:
        messagebox.showinfo("No Input Found", "Please fill all the information")
    else:
        try:
            # getting input
            x = list(map(int, entry1.get().split(",")))
            y = list(map(int, entry2.get().split(",")))
        except:
            messagebox.showinfo("Invalid Input", "please enter comma seperated Integers")
        else:
            # setting variables
            type = "bar"
            xlabel = entry3.get()
            ylabel = entry4.get()
            title = entryt.get()
            param0 = """select title from plots;"""
            titles = conn.execute(param0)
            titles_lis = []
            for r in titles:
                titles_lis.append(r[0])
            if title not in titles_lis:
                param = """insert into plots(title,type,x_values,y_values,x_label,y_label) values(?,?,?,?,?,?);"""
                columns = (title, type, entry1.get(), entry2.get(), xlabel, ylabel)
                cursor.execute(param, columns)
                conn.commit()

                fig = Figure(figsize=(10,4), dpi=100)
                plot1 = fig.add_subplot(111)
                plot1.set_xlabel(entry3.get())
                plot1.set_ylabel(entry4.get())
                plot1.bar(x, y)
                name = "bar" + title + ".png"
                fig.savefig("C:/Users/Hemanth.M/Desktop/project images/" + name)
                canvas = FigureCanvasTkAgg(fig, master=panel)
                canvas.draw()
                canvas.get_tk_widget().pack()

                path = "C:/Users/Hemanth.M/Desktop/project images/" + name
                param = """update plots set plot=? where title=?;"""
                column = (path, title)

                cursor.execute(param, column)
                conn.commit()
            else:
                messagebox.showinfo("Error", "Please enter Unique Title")


def barplt(destroyed=False):
    # destroying old window
    global root
    if not destroyed:
        root.destroy()

    # creating new window
    global r1
    r1 = Tk()
    r1.geometry("1920x1080")
    # adding frame to window
    frame_2 = LabelFrame(r1)
    frame_2.place(width=1800, height=480)
    frame_2.pack()

    # creating labels for input fields
    label1 = Label(frame_2, text="Enter X Values:", padx=20, font=(" Maiandra GD", 15))
    label2 = Label(frame_2, text="Enter Y Values:", padx=20, font=(" Maiandra GD", 15))
    label3 = Label(frame_2, text="X Label: ", padx=20, font=(" Maiandra GD", 15))
    label4 = Label(frame_2, text="Y Label:", padx=20, font=(" Maiandra GD", 15))
    title = Label(frame_2, text="Title:", padx=20, font=(" Maiandra GD", 15))

    # creating input fields
    global entry1, entry2, entry3, entry4, entryt
    entry1 = Entry(frame_2, width=30, font=(" Maiandra GD", 15))
    entry2 = Entry(frame_2, width=30, font=(" Maiandra GD", 15))
    entry3 = Entry(frame_2, width=15, font=(" Maiandra GD", 15))
    entry4 = Entry(frame_2, width=15, font=(" Maiandra GD", 15))
    entryt = Entry(frame_2, width=30, font=(" Maiandra GD", 15))

    # creating submit button to frame
    submit = Button(frame_2, text="PLOT IT", bg="green", fg="white", width=20, height=3, command=plotbar)

    # creating back button to window
    back = Button(frame_2, text="<- BACK", height=2, width=12, command=new, font=("Consolas", 10), bg="grey",
                  fg="white")
    back.grid(row=3, column=0)
    #creating select file button
    s_file = Button(frame_2,text ="selct from file",bg= "blue",fg= "white",width =20 ,height = 3,command =open_fileb)
    

    # creating reset button to frame
    reset = Button(frame_2, text="⟲", font=("Consolas", 20), width=8, bg="#DC8474", fg="white", command=barReset)
    reset.grid(row=3, column=2)

    # adding text fields,labels and buttons in order
    title.grid(row=0, column=0)
    entryt.grid(row=0, column=1)

    label1.grid(row=1, column=0)
    entry1.grid(row=1, column=1, padx=20, pady=10)

    label3.grid(row=1, column=2)
    entry3.grid(row=1, column=3, padx=10)

    label2.grid(row=2, column=0)
    entry2.grid(row=2, column=1, padx=20, pady=10)

    label4.grid(row=2, column=2)
    entry4.grid(row=2, column=3, padx=10)

    submit.grid(row=3, column=1, pady=20)
    s_file.grid(row=3,column=3)

    r1.mainloop()


def searchplot():
    global ent, hist_window
    hist_window = Tk()
    # hist_window.geometry("1920x1080")
    hist_window.state("zoomed")

    panel = LabelFrame(hist_window, pady=15, padx=10)
    panel.place(width=1800, height=480)
    panel.pack()

    back = Button(panel, text="<- BACK", command=new1, font=("Consolas", 12), bg="grey",
                  fg="white")

    lab = Label(panel, text="Enter title of the plot :", font=("consolas", 15))
    ent = Entry(panel, width=30, font=("consolas", 15))
    lab.grid(row=0, column=0)
    ent.grid(row=0, column=1)
    b = Button(panel, text="Get Plot", font=("consolas", 12), bg="green", fg="White", command=runAnother)
    b.grid(row=1, column=1)
    back.grid(row=1, column=0)

    hist_window.mainloop()


def runAnother():
    global img1, xvalues, yvalues, pievalues, pielabels
    title = ent.get()
    param0 = """select title from plots;"""
    path1 = conn.execute(param0)
    list1 = []
    for i in path1:
        list1.append(i[0])
    param1 = """select title from pieplots;"""
    path2 = conn.execute(param1)
    list2 = []
    for i in path2:
        list2.append(i[0])

    if title in list1:
        param = """select plot, x_values,y_values from plots where title=?;"""
        columns = (title,)
        path = cursor.execute(param, columns)
        conn.commit()
        img1 = ""
        for r in path:
            img1 = r[0]
            xvalues = r[1]
            yvalues = r[2]

        panel2 = LabelFrame(hist_window, padx=15, pady=15)
        panel2.pack()
        # Creates a Tkinter-compatible photo image, which can be used everywhere Tkinter expects an image object.

        img = ImageTk.PhotoImage(Image.open(img1), master=panel2)

        # The Label widget is a standard Tkinter widget used to display a text or image on the screen.
        label1 = Label(panel2, image=img, height=630)

        label1.photo = img
        # The Pack geometry manager packs widgets in rows or columns.
        label1.grid(row=2, column=0, columnspan=2)

        label2 = Label(panel2, text="X Values :", font=("Consolas", 12))
        label3 = Label(panel2, text=xvalues, font=("Consolas", 12))
        label4 = Label(panel2, text="Labels:", font=("Consolas", 12))
        label5 = Label(panel2, text=yvalues, font=("Consolas", 12))

        label2.grid(row=0, column=0)
        label3.grid(row=0, column=1)
        label4.grid(row=1, column=0)
        label5.grid(row=1, column=1)

    elif title in list2:
        param2 = """select plot, pie_values, Labels from pieplots where title=?;"""
        columns = (title,)
        path3 = cursor.execute(param2, columns)
        conn.commit()
        img1 = ""
        for r in path3:
            img1 = r[0]
            pievalues = r[1]
            pielabels = r[2]

        panel2 = LabelFrame(hist_window, padx=15, pady=15)
        panel2.pack()
        # Creates a Tkinter-compatible photo image, which can be used everywhere Tkinter expects an image object.

        img = ImageTk.PhotoImage(Image.open(img1), master=panel2)

        # The Label widget is a standard Tkinter widget used to display a text or image on the screen.
        label1 = Label(panel2, image=img, height=630)

        label1.photo = img
        # The Pack geometry manager packs widgets in rows or columns.
        label1.grid(row=2, column=0, columnspan=2)

        label2 = Label(panel2, text="pie Values :", font=("Consolas", 15))
        label3 = Label(panel2, text=pievalues, font=("Consolas", 15))
        label4 = Label(panel2, text="Labels :", font=("Consolas", 15))
        label5 = Label(panel2, text=pielabels, font=("Consolas", 15))

        label2.grid(row=0, column=0)
        label3.grid(row=0, column=1)
        label4.grid(row=1, column=0)
        label5.grid(row=1, column=1)

    else:
        messagebox.showinfo("No Title", "Entered title is not Found!")


def new1():
    hist_window.destroy()


def new():
    r1.destroy()
    main()
    
def main():
    global root
    root = Tk()
    root.state("zoomed")  # for full screen

    global conn, cursor
    conn = sqlite3.connect("C:/Users/Hemanth.M/Desktop/db/mydb.db")
    cursor = conn.cursor()

    # creating a label for adding title
    title = Label(root, text="Data Visualizer")
    title.config(font=("Maiandra GD", 44))
    title.grid(row=0, column=1)

    # creating button for checking previous plots
    history = Button(root, text="History", width=8, height=2, font=("Maiandra GD", 12), fg="white", bg="grey",
                     command=searchplot)
    history.grid(row=0, column=0)

    # creating a frame to add buttons
    frame = LabelFrame(root, text="Plot Options", padx=45, pady=70)
    frame.grid(row=1, columnspan=3, padx=10, pady=10)

    # creating Image buttons
    scatter_button = Button(frame, command=scatterplt)
    histogram_button = Button(frame, command=histplt)
    pie_button = Button(frame, command=pieplt)
    line_button = Button(frame, command=lineplt)
    bar_button = Button(frame, command=barplt)

    # creating  photos for buttons
    scatter_photo = PhotoImage(file="C:/Users/Hemanth.M/Desktop/button images/scatter.png")
    histogram_photo = PhotoImage(file="C:/Users/Hemanth.M/Desktop/button images/hist.png")
    pie_photo = PhotoImage(file="C:/Users/Hemanth.M/Desktop/button images/pie.png")
    line_photo = PhotoImage(file="C:/Users/Hemanth.M/Desktop/button images/line.png")
    bar_photo = PhotoImage(file="C:/Users/Hemanth.M/Desktop/button images/bar.png")

    # adding photos to buttons
    scatter_button.config(image=scatter_photo, width=245, height=180)
    histogram_button.config(image=histogram_photo, width=245, height=180)
    pie_button.config(image=pie_photo, width=245, height=180)
    line_button.config(image=line_photo, width=245, height=180)
    bar_button.config(image=bar_photo, width=245, height=180)

    # adding buttons to grid
    scatter_button.grid(row=0, column=0)
    histogram_button.grid(row=0, column=2)
    pie_button.grid(row=0, column=4)
    line_button.grid(row=1, column=1)
    bar_button.grid(row=1, column=3)

    root.mainloop()

main()
